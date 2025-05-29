import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from PIL import Image
import io, urllib.parse, requests, tempfile, math, unicodedata

# Configuración general
st.set_page_config(
    page_title="Catálogo Millex",
    page_icon="🐾",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={"Get Help": None, "Report a bug": None, "About": None},
)

# CSS global (oculta logos, estilos, FAB, botón cerrar carrito)
st.markdown(
    """
<style>
/* Ocultar menús / logos */
#MainMenu, footer, header {visibility: hidden;}
.viewerBadge_container__1QSob,
.viewerBadge_container__rGiy7,
a[href="https://streamlit.io"],
div[class^="viewerBadge_container"],
.stDeployButton {display:none!important;}

/* Ajuste top padding */
.block-container {padding-top:1rem;}

/* Nuevas reglas para móvil */
@media(max-width:768px){
  /* Paginación móvil - flechas juntas */
  .pagination-mobile{display:flex;justify-content:center;gap:16px;margin:20px 0;}
  .pagination-mobile button{background:#f0f2f6;border:none;border-radius:6px;
    padding:8px 16px;cursor:pointer;transition:.3s;font-size:18px;}
  .pagination-mobile button:hover{background:#e0e2e6;}
  .pagination-mobile button:disabled{opacity:.5;cursor:not-allowed;}

  /* Ocultar paginación normal en móvil */
  .pagination{display:none;}

  /* Mostrar paginación móvil */
  .mobile-pager{display:block!important;}

  /* Reducir productos por página en móvil */
  .mobile-items-per-page{display:block!important;}

  /* Ocultar botón de carrito desktop en móvil, ya que tenemos el FAB */
  .desktop-cart-button-container {
      display:none!important;
  }
}

/* Ocultar paginación móvil en desktop */
.mobile-pager{display:none;}
.mobile-items-per-page{display:none;}

/* FAB carrito (solo mobile) */
.carrito-fab{
  position:fixed;bottom:16px;right:16px;
  background:#f63366;color:#fff;
  padding:14px 20px;font-size:18px;font-weight:700;
  border-radius:32px;box-shadow:0 4px 12px rgba(0,0,0,.35);
  z-index:99999;cursor:pointer;transition:transform .15s;
  display:flex;align-items:center;justify-content:center;gap:8px;
}
.carrito-fab:hover{transform:scale(1.06);}
@media(min-width:769px){.carrito-fab{display:none;}}/* solo cel/tablet - OJO: Si quieres el FAB en desktop, comenta o borra esta línea */

/* Botón Carrito Desktop (Opción 2) */
.desktop-cart-button-container {
    display: flex;
    align-items: flex-end; /* Alinea el botón con la base de los inputs */
    height: 100%;
    padding-bottom: 0px; /* Ajustado para que el botón se alinee mejor con st.text_input */
}
/* Estilo para el botón HTML para que se parezca a los de Streamlit pero con colores personalizados */
.desktop-cart-button-container button.custom-st-button {
    width: 100%;
    background-color: #f63366;
    color: white;
    border: none;
    padding: 0.5rem 1rem; /* Similar al padding de st.button */
    border-radius: 0.5rem; /* Similar al border-radius de st.button */
    font-weight: 600; /* Similar a st.button */
    font-size: 0.875rem; /* Similar a st.button */
    cursor: pointer;
    transition: background-color 0.3s ease;
    height: 40px; /* Ajustar altura para que coincida con st.text_input */
    line-height: 24px; /* Ajustar para centrar texto verticalmente */
}
.desktop-cart-button-container button.custom-st-button:hover {
    background-color: #e02b5a;
    color: white;
}
.desktop-cart-button-container button.custom-st-button:focus {
    outline: none;
    box-shadow: 0 0 0 2px rgba(246, 51, 102, 0.5); /* Sombra de foco similar a Streamlit */
}

/* Productos */
.product-card{border:1px solid #e0e0e0;border-radius:12px;
  padding:16px;height:100%;transition:box-shadow .3s;
  display:flex;flex-direction:column;}
.product-card:hover{box-shadow:0 4px 12px rgba(0,0,0,.1);}
.product-image{width:100%;height:180px;object-fit:contain;
  margin-bottom:12px;border-radius:8px;background:#f9f9f9;}
.product-title{font-size:16px;font-weight:600;margin-bottom:8px;color:#333;flex-grow:1;}
.product-code{font-size:14px;color:#666;margin-bottom:4px;}
.product-price{font-size:18px;font-weight:700;color:#f63366;margin-bottom:12px;}
.stNumberInput>div,.stNumberInput input{width:100%;}

/* Paginación */
.pagination{display:flex;justify-content:center;margin:20px 0;gap:8px;}
.pagination button{background:#f0f2f6;border:none;border-radius:6px;
  padding:8px 12px;cursor:pointer;transition:.3s;}
.pagination button:hover{background:#e0e2e6;}
.pagination button.active{background:#f63366;color:#fff;}
.pagination button:disabled{opacity:.5;cursor:not-allowed;}

/* Sidebar (carrito) */
[data-testid="stSidebar"]{background:#f8f9fa;padding:16px;position:relative;}
.sidebar-title{display:flex;align-items:center;gap:8px;margin-bottom:16px;}
.cart-item{padding:12px 0;border-bottom:1px solid #e0e0e0;color:#333;}
.cart-item:last-child{border-bottom:none;}
.cart-total{font-weight:700;font-size:18px;margin:16px 0;color:#f63366;}
.close-sidebar{position:absolute;top:10px;right:14px;font-size:22px;
  cursor:pointer;color:#666;user-select:none;}
.close-sidebar:hover{color:#000;}
.whatsapp-btn{background:#25D366!important;color:#fff!important;width:100%;margin:8px 0;}
.clear-btn{background:#f8f9fa!important;color:#f63366!important;
  border:1px solid #f63366!important;width:100%;margin:8px 0;}
</style>
""",
    unsafe_allow_html=True,
)

# Utilidades
def quitar_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(texto))
        if not unicodedata.combining(c)
    ).lower()

# Descarga del Excel (cacheado)
@st.cache_data(show_spinner=False)
def fetch_excel(file_id: str) -> Path:
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    tmp = Path(tempfile.gettempdir()) / f"{file_id}.xlsx"
    tmp.write_bytes(r.content)
    return tmp

# Lectura de productos + imágenes (cacheado)
@st.cache_data(show_spinner=False)
def load_products(xls_path: str) -> pd.DataFrame:
    wb = load_workbook(xls_path, data_only=True)
    ws = wb.active
    img_map = {img.anchor._from.row + 1: img._data() for img in ws._images if hasattr(img, "_data")}
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row[1]:  # columna B vacía => fin
            break
        codigo, detalle, precio = row[1], row[2], row[3]
        precio = 0 if precio is None else float(str(precio).replace("$", "").replace(",", ""))
        rows.append({"fila_excel": idx, "codigo": str(codigo), "detalle": str(detalle), "precio": precio})
    df = pd.DataFrame(rows)
    df["img_bytes"] = df["fila_excel"].map(img_map)
    # Columnas normalizadas para búsqueda
    df["codigo_norm"]  = df["codigo"].apply(quitar_acentos)
    df["detalle_norm"] = df["detalle"].apply(quitar_acentos)
    return df

# IDs de hojas (líneas de productos)
FILE_IDS = {
    "Línea Perros": "1EK_NlWT-eS5_7P2kWwBHsui2tKu5t26U",
    "Línea Pájaros y Roedores": "1n10EZZvZq-3M2t3rrtmvW7gfeB40VJ7F",
    "Línea Gatos": "1vSWXZKsIOqpy2wNhWsKH3Lp77JnRNKbA",
    "Línea Bombas de Acuario": "1DiXE5InuxMjZio6HD1nkwtQZe8vaGcSh",
}

# UI: selector de línea + buscador + botón carrito (Opción 2)
col_linea, col_search, col_cart_btn_placeholder = st.columns([2.2, 3, 1.5]) # Ajustar ratios según necesidad
with col_linea:
    linea = st.selectbox("Elegí la línea de productos:", list(FILE_IDS.keys()), label_visibility="collapsed", placeholder="Elegí la línea de productos:")
with col_search:
    search_term = st.text_input("🔍 Buscar (código o descripción)…", placeholder="🔍 Buscar (código o descripción)…", label_visibility="collapsed").strip().lower()
search_norm = quitar_acentos(search_term)

# Botón para abrir el carrito en desktop (se oculta en móvil si el FAB ya existe)
with col_cart_btn_placeholder:
    qty_total_header = sum(it["qty"] for it in st.session_state.get("cart", {}).values())
    cart_btn_label = f"🛒 Carrito ({qty_total_header})" if qty_total_header else "🛒 Ver Carrito"

    # Usamos HTML para poder aplicar una clase y ocultarlo en móvil, y para el onclick JS
    st.markdown(
        f"""
        <div class="desktop-cart-button-container">
            <button class="custom-st-button" onclick="window.dispatchEvent(new Event('toggleSidebar'))">{cart_btn_label}</button>
        </div>
        """,
        unsafe_allow_html=True
    )

# Carga y filtrado del catálogo
df_base = load_products(str(fetch_excel(FILE_IDS[linea])))

if search_term:
    df = df_base[
        df_base["codigo_norm"].str.contains(search_norm, na=False)
        | df_base["detalle_norm"].str.contains(search_norm, na=False)
    ]
else:
    df = df_base.copy()

# Paginación
ITEMS_PER_PAGE_DESKTOP = 45
ITEMS_PER_PAGE_MOBILE = 10

# La CSS ya se encarga de mostrar/ocultar elementos de paginación y ajustar visualización
ITEMS_PER_PAGE = ITEMS_PER_PAGE_DESKTOP  # Usaremos el de desktop para el cálculo de páginas

total_pages = max(1, math.ceil(len(df) / ITEMS_PER_PAGE))
page_key = f"current_page_{linea}_{search_term}"  # Añadimos search_term para resetear página en nueva búsqueda
if page_key not in st.session_state:
    st.session_state[page_key] = 1
current_page = min(st.session_state.get(page_key, 1), total_pages)

def change_page(new_page_val: int):
    st.session_state[page_key] = new_page_val

def pager(position: str):
    # Versión móvil (flechas juntas) - controlada por CSS
    st.markdown(
        f"""
    <div class="mobile-pager">
      <div class="pagination-mobile">
        <button onclick="window.dispatchEvent(new CustomEvent('streamlit_page_change', {{detail: {{page: {current_page - 1}, position: '{position}', direction: 'prev' }} }}))"
                {'disabled' if current_page == 1 else ''}>◀</button>
        <span style="padding:8px 12px;font-weight:bold;">Pág. {current_page}/{total_pages}</span>
        <button onclick="window.dispatchEvent(new CustomEvent('streamlit_page_change', {{detail: {{page: {current_page + 1}, position: '{position}', direction: 'next' }} }}))"
                {'disabled' if current_page == total_pages else ''}>▶</button>
      </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # Versión desktop (original) - controlada por CSS
    st.markdown('<div class="pagination">', unsafe_allow_html=True)

    cols_pager = st.columns([1, 1, 1])  # Dividimos en 3 para los botones y el texto

    with cols_pager[0]:
        if st.button("◀ Anterior", key=f"{position}_prev_desktop", disabled=current_page == 1, use_container_width=True):
            change_page(current_page - 1)
            st.rerun()  # Forzar rerun para actualizar la vista con la nueva página

    with cols_pager[1]:
        st.markdown(f"<div style='text-align: center; padding: 0.25rem;'>Página {current_page} de {total_pages}</div>", unsafe_allow_html=True)

    with cols_pager[2]:
        if st.button("Siguiente ▶", key=f"{position}_next_desktop", disabled=current_page == total_pages, use_container_width=True):
            change_page(current_page + 1)
            st.rerun()  # Forzar rerun

    st.markdown('</div>', unsafe_allow_html=True)

# JS para manejar los eventos de los botones de paginación móvil (HTML)
st.markdown(
    """
<script>
document.addEventListener('streamlit_page_change', function(event) {
    const { page, position, direction } = event.detail;
    let buttonToClick;

    if (direction === 'prev') {
        buttonToClick = window.parent.document.querySelectorAll('button[data-testid="stButton"] > div > p:contains("◀ Anterior")')[0];
        if (!buttonToClick) {
            const buttons = window.parent.document.querySelectorAll('button[data-testid="stButton"]');
            for (let i = 0; i < buttons.length; i++) {
                const keyAttr = buttons[i].getAttribute('key');
                if (keyAttr && keyAttr.includes(position + '_prev_desktop')) {
                    buttonToClick = buttons[i];
                    break;
                }
            }
        }
    } else if (direction === 'next') {
        buttonToClick = window.parent.document.querySelectorAll('button[data-testid="stButton"] > div > p:contains("Siguiente ▶")')[0];
        if (!buttonToClick) {
            const buttons = window.parent.document.querySelectorAll('button[data-testid="stButton"]');
            for (let i = 0; i < buttons.length; i++) {
                const keyAttr = buttons[i].getAttribute('key');
                if (keyAttr && keyAttr.includes(position + '_next_desktop')) {
                    buttonToClick = buttons[i];
                    break;
                }
            }
        }
    }

    if (buttonToClick) {
        buttonToClick.click();
    } else {
        console.warn("Mobile pagination button couldn't find corresponding Streamlit button for position: " + position + ", direction: " + direction);
    }
});
</script>
""",
    unsafe_allow_html=True,
)

if total_pages > 1:
    pager("top")

# Mostrar productos (grilla 3xN)
start_idx = (current_page - 1) * ITEMS_PER_PAGE
end_idx = current_page * ITEMS_PER_PAGE
paginated_df = df.iloc[start_idx:end_idx]

if paginated_df.empty and len(df) > 0:  # Si la página actual está vacía pero hay datos (ej. se borró de otra página)
    st.session_state[page_key] = 1  # Volver a la página 1
    st.rerun()
elif paginated_df.empty and search_term:
    st.info("No se encontraron productos que coincidan con tu búsqueda.")
elif paginated_df.empty:
    st.info("No hay productos para mostrar en esta línea.")

for i in range(0, len(paginated_df), 3):
    cols = st.columns(3)
    for j in range(3):
        if i + j >= len(paginated_df):
            with cols[j]:  # Dejar la columna vacía si no hay producto
                st.container()
            continue
        prod = paginated_df.iloc[i + j]
        with cols[j]:
            st.markdown('<div class="product-card">', unsafe_allow_html=True)

            # Imagen
            if pd.notna(prod.img_bytes) and len(prod.img_bytes) > 0:
                try:
                    st.image(Image.open(io.BytesIO(prod.img_bytes)), use_container_width=True, output_format='PNG')
                except Exception as e:
                    st.image("https://via.placeholder.com/200x150?text=Error+Img", use_container_width=True)
            else:
                st.image("https://via.placeholder.com/200x150?text=Sin+imagen", use_container_width=True)

            # Texto
            st.markdown(f'<div class="product-title">{prod.detalle}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="product-code">Código: {prod.codigo}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="product-price">${prod.precio:,.2f}</div>', unsafe_allow_html=True)

            # Selector cantidad
            qty_key = f"qty_{linea}_{prod.codigo}"

            # Recuperar la cantidad del carrito para este producto
            cart = st.session_state.setdefault("cart", {})
            current_qty_in_cart = cart.get(str(prod.codigo), {}).get("qty", 0)

            qty = st.number_input("Cantidad", min_value=0, step=1,
                                  key=qty_key,
                                  value=current_qty_in_cart)  # El valor inicial es el del carrito

            # Actualizar Carrito en sesión si la cantidad cambia
            if qty != current_qty_in_cart:  # Solo actualizar si hay un cambio
                if qty > 0:
                    cart[str(prod.codigo)] = {"detalle": prod.detalle, "precio": prod.precio, "qty": qty, "linea": linea}
                elif str(prod.codigo) in cart:  # Si la cantidad es 0 y estaba en el carrito, eliminarlo
                    del cart[str(prod.codigo)]
                st.rerun()  # Necesario para actualizar el contador del botón del carrito y el FAB

            st.markdown("</div>", unsafe_allow_html=True)

# Paginador inferior
if total_pages > 1:
    pager("bottom")

# Sidebar ➜ Carrito
with st.sidebar:
    st.markdown('<div class="close-sidebar" onclick="window.dispatchEvent(new Event(\'toggleSidebar\'))">✖</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title"><h2>🛒 Carrito</h2></div>', unsafe_allow_html=True)
    st.markdown("---")

    cart = st.session_state.get("cart", {}) # Usar .get para seguridad
    if cart:
        for cod, it in cart.items():
            st.markdown(
                f"""
<div class="cart-item">
  <div><strong>{it['detalle']}</strong></div>
  <div>Código: {cod}</div>
  <div>Cantidad: {it['qty']}</div>
  <div>Subtotal: ${it['precio'] * it['qty']:,.2f}</div>
</div>
""",
                unsafe_allow_html=True,
            )

        total = sum(it["precio"] * it["qty"] for it in cart.values())
        st.markdown(f'<div class="cart-total">Total: ${total:,.2f}</div>', unsafe_allow_html=True)

        # Enlace WhatsApp
        msg_lines = [f"- {it['detalle']} (Código {cod}) x {it['qty']}" for cod, it in cart.items()]
        msg = "Hola! Quiero hacer un pedido de los siguientes productos:\n" + "\n".join(msg_lines) + f"\n\nTotal: ${total:,.2f}"
        link = f"https://wa.me/5493516434765?text={urllib.parse.quote(msg)}"

        # Usar st.link_button para el botón de WhatsApp
        st.link_button("📲 Confirmar pedido por WhatsApp", link, use_container_width=True, type="primary")

        if st.button("🗑️ Vaciar carrito", key="clear_btn_sidebar", use_container_width=True, type="secondary"):
            # Crear una copia de las claves de los productos en el carrito antes de limpiarlo
            keys_to_reset = []
            for product_code_in_cart, item_details in cart.items():
                original_linea = item_details.get("linea", linea) # Fallback a la línea actual si no se guardó
                keys_to_reset.append(f"qty_{original_linea}_{product_code_in_cart}")

            cart.clear() # Limpiar el carrito

            # Resetear los st.number_input a 0
            for k_to_reset in keys_to_reset:
                if k_to_reset in st.session_state:
                    st.session_state[k_to_reset] = 0

            st.rerun()
    else:
        st.write("Todavía no agregaste productos.")

# FAB móvil (actualizar su etiqueta)
qty_total_fab = sum(it["qty"] for it in st.session_state.get("cart", {}).values())
fab_label = f"🛒 ({qty_total_fab})" if qty_total_fab else "🛒 Ver carrito"
st.markdown(
    f'<div class="carrito-fab" onclick="window.dispatchEvent(new Event(\'toggleSidebar\'))">{fab_label}</div>',
    unsafe_allow_html=True,
)

# JS global: alternar sidebar
st.markdown(
    """
<script>
window.addEventListener("toggleSidebar", () => {
  const btn = window.parent.document.querySelector('button[data-testid="stSidebarNavToggler"]');
  if (btn) {
    btn.click();
  } else {
    const olderBtn = window.parent.document.querySelector('button[aria-label^="Toggle sidebar"]') ||
                     window.parent.document.querySelector('button[title^="Expand sidebar"]') ||
                     window.parent.document.querySelector('button[title^="Collapse sidebar"]');
    if (olderBtn) olderBtn.click();
    else console.warn("Sidebar toggle button not found.");
  }
});
</script>
""",
    unsafe_allow_html=True,
)


